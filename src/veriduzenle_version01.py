#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Aug 26 16:44:12 2023

@author: senolirmak
"""

import openpyxl
import pandas as pd
# gereksiz uyarıları kapatalım
import warnings
warnings.filterwarnings('ignore')


class veriDuzenle:
    def __init__(self, dosya_xls):
        self.dataframe = openpyxl.load_workbook(dosya_xls)
        self.dataSheet = self.dataframe.active
        self.max_rov = self.dataSheet.max_row
        self.duzenlidata = []

    @staticmethod
    def baslik_temizle(baslik):
        temizle = ['Sınıf', 'Şubesi', '.', '-', '/', '(', ')']
        t_sube = []
        for terim in temizle:
            if terim == '-':
                baslik = baslik.replace(terim, " ")
            baslik = baslik.replace(terim, "")
        t_sube = baslik.split()
        t_sube.pop()
        alanAdi = ' '.join(t_sube[3:])
        b_sube = t_sube[:3]
        b_sube.append(alanAdi)
        return b_sube

    
    def sinif_veri(self, baslik):

        t_sube = self.baslik_temizle(baslik)
        alan_ad = ' '.join(t_sube[3:]).split()[0]
        return {
            'tur': t_sube[0],
            'sinif': t_sube[1],
            'sube': t_sube[2],
            'alan': t_sube[3],
            'alan_ad': alan_ad
        }
        
    def aralik(self, r0 = 0):
        ara = ["AMP","ATP"]
        for row in range(r0, self.max_rov):
            for col in self.dataSheet.iter_cols(0, 0):
                if str(col[row].value).find(ara[0]) == 0 or str(col[row].value).find(ara[1]) == 0 :
                    return row , self.sinif_veri(str(col[row].value))
              
                    
        return None, None
    
    
    def sorumluDersVerileri(self):
        r0 = 0
        duzenlidata = []
        veribaslik = ['duzey', 'ders','okulno','adisoyadi','okul', 
                      'sinif','sube','alan']
        veri = dict.fromkeys(veribaslik)
        no_adi = {'okulno':"",'adisoyadi':""}
        
        
        while r0 < (self.max_rov):
            satir_ust , sinif = self.aralik(r0)
            
            k = 0
            n = 0
            if not (satir_ust is None):
                satir_alt , veri_alt = self.aralik(satir_ust+1)
                if satir_alt is None:
                    satir_alt = self.max_rov
                    
                for row in range(satir_ust+2 , satir_alt):
                    for col in self.dataSheet.iter_cols(9, 10):
                        if col[row].value is not None:
                            veri[list(veri.keys())[k]] = col[row].value
                            k += 1
                    k = 0        
                            
                    for col in self.dataSheet.iter_cols(2, 3):
                        if col[row].value is not None:
                            no_adi[list(no_adi.keys())[n]] = col[row].value
                            n += 1
                    
                    n = 0
                        
                    veri['okul'] = sinif['tur']
                    veri['alan'] = sinif['alan']
                    veri['sinif'] = sinif['sinif']
                    veri['sube'] = sinif['sube']
                    veri['okulno'] = no_adi['okulno']
                    veri['adisoyadi'] = no_adi['adisoyadi']
                    duzenlidata.append(veri)
                    veri = dict.fromkeys(veribaslik)
                r0 = satir_ust+1
            else:
                break
            
        p_data = pd.DataFrame(duzenlidata)
        p_data = p_data.dropna()
        p_data = p_data.reset_index(drop=True)
        return p_data
    
    def dersEki(self, sayi):
        sayi = int(sayi)
        ek = str(sayi) if len(str(sayi))== 2 else "0" + str(sayi)
        return ek
    
    def sorumluDersData(self):

        sddata = []
        veri = {}

        ogrSorumluDers = self.sorumluDersVerileri()
        _, sorumludersArry = pd.factorize(ogrSorumluDers['ders'])
        sorumludersList = sorumludersArry.tolist()
        newdf = ogrSorumluDers[['duzey','ders']]
        aa= newdf.to_dict('records')
        
        for count, ders in enumerate(sorumludersList):
            for idx, rec1 in enumerate(aa):
                if rec1['ders'] == ders :
                    kod = "ders_"+self.dersEki(count)+self.dersEki(rec1['duzey'])
                    veri ['row'] = idx
                    veri ['derskodu'] = kod
                    zz = veri.copy()
                    sddata.append(zz)
                    veri.clear()
        
        sorted_dizi_reverse = sorted(sddata, key=lambda x: x['row'], reverse=False)
        ogrSorumluDers['derskodu'] = [ x['derskodu'] for x in sorted_dizi_reverse]
        
        derskodlari = ogrSorumluDers[['duzey','ders','derskodu']]
        kodDers = derskodlari.drop_duplicates(subset=['derskodu'])
        _, kodArry = pd.factorize(kodDers['derskodu'])
        kodList = kodArry.tolist()
        ogrSorumluDers = pd.concat([ogrSorumluDers, pd.DataFrame(columns = kodList)])
        _, ogrenciNoArry = pd.factorize(ogrSorumluDers['okulno'])
        ogrenciNoList = ogrenciNoArry.tolist()
        
        for ogrenciNo in ogrenciNoList :
            nwkisi = ogrSorumluDers.loc[ogrSorumluDers['okulno'] == ogrenciNo]
            nwkodu = nwkisi['derskodu'].tolist()
            indexs = nwkisi.index[nwkisi['okulno'] == ogrenciNo].tolist()
            for colname in nwkodu:
                ogrSorumluDers.at[indexs[0],colname] = 1
            
        ogrSorumluDers = ogrSorumluDers.drop_duplicates(subset=['okulno'])
        return ogrSorumluDers, kodDers
    
    def degistirAlan(self, _list1):
        _list = [['TEKNOLOJİLERİ',''],
                 ['TEKNOLOJİSİ',''],
                 ['MAKİNE VE TASARIM','MAKİNE'],
                 ['MOBİLYA VE İÇ MEKÂN TASARIMI','MOBİLYA'],
                 ['ELEKTRİK ELEKTRONİK','ELEKTRİK']]

        sub = dict(_list)
        for key, val in sub.items():
            for idx, ele in enumerate(_list1):
                if key in ele:
                    _list1[idx] = ele.replace(key, val).strip()
        
        return _list1
    
    def dersKoduAlanBelirle(self,dersmatrisdata,derskod):
        
        derskod.reset_index(drop=True, inplace=True)
        indexs = derskod.index.tolist()    
        derskod['alan'] = None

        for i in indexs:
            aa = ["-".join(pd.Series(dersmatrisdata.query(derskod.loc[i, ['derskodu']].values[0] +" == 1")['alan'].tolist()).drop_duplicates().tolist())]
            bb = self.degistirAlan(aa)
            derskod.loc[i, ['alan']] = bb
        
        derskod.reset_index(drop=True, inplace=True)
        return derskod
        
    def dataFrametoExcel(self, donem:str):
        sorumluDersMatris, sorumluDersKod = self.sorumluDersData()
        sorumluDersMatris.drop(sorumluDersMatris.columns[[0,1, 8]], axis=1, inplace=True)
        sorumluDersKod = self.dersKoduAlanBelirle(sorumluDersMatris, sorumluDersKod)
        deskodlari = sorumluDersKod.derskodu.values
        
        for ders in deskodlari:
            sorumluDersMatris[ders]=sorumluDersMatris[ders].fillna(0)
            sorumluDersMatris[ders]=sorumluDersMatris[ders].astype(int)
          
        sorumluDersMatris.to_excel("2023_DersCikti_"+donem+".xlsx", index=False)
        sorumluDersKod.to_excel("2023_Kod_"+donem+".xlsx", index=False)

